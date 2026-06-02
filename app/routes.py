import csv
import io
import threading
import uuid
from collections import Counter
from datetime import datetime

from flask import Blueprint, Response, current_app, flash, jsonify, redirect, render_template, request, session, url_for
from sqlalchemy import and_, case, or_

from . import db
from .analyzer import (
    analyze_sentiment,
    classify_category,
    find_existing_review,
    review_storage_id,
)
from .storage_health import storage_health_report
from .datetime_utils import (
    coerce_review_datetime as _coerce_review_datetime,
    parse_csv_review_date as _parse_csv_review_date,
)
from .app_catalog import (
    catalog_status,
    has_strong_local_match,
    is_package_query,
    load_catalog,
    lookup_local_by_package,
    search_local_catalog,
)
from .google_play import (
    fetch_google_play_reviews,
    fetch_google_play_reviews_all,
    lookup_app_by_package,
    merge_and_rank_suggestions,
    search_apps_play,
)
from .models import ProcessingLog, Review, Ticket
from .ticketing import (
    choose_ticket_platform,
    create_jira_ticket,
    create_zendesk_ticket,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

main_bp = Blueprint("main", __name__)
FETCH_JOBS = {}
FETCH_JOBS_LOCK = threading.Lock()

POPULAR_APPS = [
    {
        "app_name": "WhatsApp Messenger",
        "package_name": "com.whatsapp",
        "icon": "https://cdn.simpleicons.org/whatsapp/25D366",
    },
    {
        "app_name": "Instagram",
        "package_name": "com.instagram.android",
        "icon": "https://cdn.simpleicons.org/instagram/E4405F",
    },
    {
        "app_name": "Facebook",
        "package_name": "com.facebook.katana",
        "icon": "https://cdn.simpleicons.org/facebook/1877F2",
    },
    {
        "app_name": "TikTok",
        "package_name": "com.zhiliaoapp.musically",
        "icon": "https://cdn.simpleicons.org/tiktok/000000",
    },
    {
        "app_name": "YouTube",
        "package_name": "com.google.android.youtube",
        "icon": "https://cdn.simpleicons.org/youtube/FF0000",
    },
    {
        "app_name": "Snapchat",
        "package_name": "com.snapchat.android",
        "icon": "https://cdn.simpleicons.org/snapchat/FFFC00",
    },
]
POPULAR_APPS_BY_NAME = {item["app_name"]: item for item in POPULAR_APPS}


def log_message(message: str, level: str = "info"):
    db.session.add(ProcessingLog(message=message, level=level))
    db.session.commit()


def _normalize_batch_dt(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is not None:
        from datetime import timezone

        value = value.astimezone(timezone.utc).replace(tzinfo=None)
    return value.replace(microsecond=0)


def _batch_now() -> datetime:
    return _normalize_batch_dt(datetime.utcnow())  # type: ignore[return-value]


def _parse_dashboard_since(value: str | None) -> datetime | None:
    if not value or not str(value).strip():
        return None
    try:
        return _normalize_batch_dt(datetime.fromisoformat(str(value).strip()))
    except ValueError:
        return None


def _dashboard_view_since() -> datetime | None:
    """Dashboard only shows a batch when `?since=ISO` is present (refresh / new visit = clean)."""
    return _parse_dashboard_since(request.args.get("since"))


def _batch_review_filter(since: datetime):
    since = _normalize_batch_dt(since)
    return or_(
        Review.last_batch_at == since,
        and_(Review.last_batch_at.is_(None), Review.created_at >= since),
    )


def _batch_reviews_query(since: datetime):
    rank_first = case((Review.play_rank.is_(None), 1), else_=0)
    date_fallback = case((Review.reviewed_at.is_(None), 1), else_=0)
    return (
        Review.query.filter(_batch_review_filter(since))
        .order_by(
            rank_first,
            Review.play_rank.asc(),
            date_fallback,
            Review.reviewed_at.desc(),
            Review.id.asc(),
        )
    )


def _completion_message(
    new_count: int,
    refreshed: int,
    skipped: int,
    *,
    tickets_created: int = 0,
    tickets_skipped_positive: int = 0,
) -> str:
    parts = [f"{new_count} new"]
    if refreshed:
        parts.append(f"{refreshed} refreshed")
    if skipped:
        parts.append(f"{skipped} skipped (invalid only)")
    if tickets_created:
        parts.append(f"{tickets_created} ticket(s) created (new reviews only)")
    if tickets_skipped_positive:
        parts.append(f"{tickets_skipped_positive} positive skipped (no ticket)")
    return "Analysis complete — " + ", ".join(parts)


def _maybe_upgrade_review_id(existing: Review, preferred_id: str) -> None:
    if existing.review_id == preferred_id:
        return
    clash = Review.query.filter(
        Review.review_id == preferred_id,
        Review.id != existing.id,
    ).first()
    if clash is None:
        existing.review_id = preferred_id


def _refresh_existing_review(
    existing: Review,
    *,
    app_name: str,
    author: str,
    rating: int,
    raw_text: str,
    sentiment: str,
    category: str,
    confidence: float,
    batch_started_at: datetime,
    play_review_id: str | None,
    reviewed_at: datetime | None,
    play_rank: int | None,
) -> None:
    preferred_id = review_storage_id(app_name, play_review_id, author, raw_text, rating)
    existing.app_name = app_name
    existing.author = author
    existing.rating = rating
    existing.content = raw_text
    existing.sentiment = sentiment
    existing.category = category
    existing.confidence = round(confidence, 2)
    if reviewed_at is not None:
        existing.reviewed_at = reviewed_at
    if play_rank is not None:
        existing.play_rank = play_rank
    existing.last_batch_at = batch_started_at
    _maybe_upgrade_review_id(existing, preferred_id)


def _create_ticket_for_review(
    review: Review,
    category: str,
    sentiment: str,
    *,
    skip_positive_tickets: bool = False,
) -> str | None:
    """Create at most one ticket per review. Returns platform name or None."""
    if review.tickets:
        return None
    if skip_positive_tickets and sentiment == "positive":
        return None

    platform = choose_ticket_platform(category)
    try:
        ticket_payload = (
            create_jira_ticket(review, review.id)
            if platform == "Jira"
            else create_zendesk_ticket(review, review.id)
        )
    except Exception as e:
        log_message(f"Ticket creation failed for review {review.id}: {e}", "error")
        ticket_payload = {
            "platform": platform,
            "external_ticket_id": "FAILED",
            "title": f"Ticket failed for review {review.id}",
            "status": "error",
        }

    db.session.add(
        Ticket(
            review_id=review.id,
            platform=ticket_payload["platform"],
            external_ticket_id=ticket_payload["external_ticket_id"],
            title=ticket_payload["title"],
            status=ticket_payload.get("status", "open"),
        )
    )
    return platform


def _parse_review_count(raw, default: int = 100) -> int:
    try:
        n = int(raw)
    except (TypeError, ValueError):
        n = default
    if n < 1:
        raise ValueError("Review count must be at least 1.")
    return n


def _parse_skip_positive_tickets(raw) -> bool:
    return str(raw or "").strip().lower() in ("1", "true", "on", "yes")


def _set_job(job_id: str, **kwargs):
    with FETCH_JOBS_LOCK:
        job = FETCH_JOBS.get(job_id, {})
        job.update(kwargs)
        FETCH_JOBS[job_id] = job


def _get_job(job_id: str):
    with FETCH_JOBS_LOCK:
        return FETCH_JOBS.get(job_id)


def _pipeline_snapshot_from_job(job: dict) -> dict:
    return {
        "status": job.get("status"),
        "phase": job.get("phase"),
        "progress": job.get("progress", 0),
        "message": job.get("message", ""),
        "app_name": job.get("app_name", "App"),
        "app_icon": job.get("app_icon", ""),
        "job_type": job.get("job_type", "play_fetch"),
        "fetched": job.get("fetched", 0),
        "processed": job.get("processed", 0),
        "new": job.get("new", 0),
        "refreshed": job.get("refreshed", 0),
        "skipped": job.get("skipped", 0),
        "jira_tickets": job.get("jira_tickets", 0),
        "zendesk_tickets": job.get("zendesk_tickets", 0),
        "total_reviews": job.get("total_reviews", 0),
    }


def _process_review(
    app_name: str,
    author: str,
    rating: int,
    text: str,
    *,
    batch_started_at: datetime,
    play_review_id: str | None = None,
    reviewed_at: datetime | None = None,
    play_rank: int | None = None,
    skip_positive_tickets: bool = False,
) -> tuple[bool, str, dict]:
    batch_started_at = _normalize_batch_dt(batch_started_at)  # type: ignore[assignment]
    raw_text = (text or "").strip()
    if not raw_text:
        return False, "empty", {}

    existing = find_existing_review(app_name, play_review_id, author, raw_text, rating)

    sentiment, confidence = analyze_sentiment(raw_text)
    category = classify_category(raw_text, rating, sentiment)

    if existing:
        _refresh_existing_review(
            existing,
            app_name=app_name,
            author=author,
            rating=rating,
            raw_text=raw_text,
            sentiment=sentiment,
            category=category,
            confidence=confidence,
            batch_started_at=batch_started_at,
            play_review_id=play_review_id,
            reviewed_at=reviewed_at,
            play_rank=play_rank,
        )
        return True, "refreshed", {
            "platform": None,
            "category": category,
            "sentiment": sentiment,
        }

    storage_id = review_storage_id(app_name, play_review_id, author, raw_text, rating)
    review = Review(
        source="Google Play",
        app_name=app_name,
        review_id=storage_id,
        author=author,
        rating=rating,
        content=raw_text,
        sentiment=sentiment,
        category=category,
        confidence=round(confidence, 2),
        reviewed_at=reviewed_at,
        last_batch_at=batch_started_at,
        play_rank=play_rank,
    )

    db.session.add(review)
    db.session.flush()

    platform = _create_ticket_for_review(
        review,
        category,
        sentiment,
        skip_positive_tickets=skip_positive_tickets,
    )
    return True, "processed", {
        "platform": platform,
        "category": category,
        "sentiment": sentiment,
    }


def _parse_review_row(row: dict, *, default_rank: int) -> dict:
    reviewed_at = _coerce_review_datetime(row.get("at") or row.get("reviewed_at"))
    if reviewed_at is None:
        reviewed_at = _parse_csv_review_date(row)
    play_review_id = (row.get("play_review_id") or "").strip() or None
    raw_rank = row.get("play_rank")
    play_rank = int(raw_rank) if raw_rank is not None else default_rank
    return {
        "author": (row.get("author") or "anonymous").strip(),
        "text": (row.get("content") or "").strip(),
        "rating": int(row.get("rating") or 3),
        "play_review_id": play_review_id,
        "reviewed_at": reviewed_at,
        "play_rank": play_rank,
    }


def _process_reviews_loop(
    job_id: str,
    app_name: str,
    rows: list,
    batch_started_at: datetime,
    app_icon: str = "",
    *,
    skip_positive_tickets: bool = False,
) -> None:
    total_rows = len(rows)
    processed = 0
    new_count = 0
    refreshed = 0
    skipped = 0
    jira_tickets = 0
    zendesk_tickets = 0
    tickets_skipped_positive = 0

    _set_job(
        job_id,
        phase="analyzing",
        progress=30,
        total_reviews=total_rows,
        fetched=total_rows,
        current_index=0,
        processed=0,
        new=0,
        refreshed=0,
        skipped=0,
        jira_tickets=0,
        zendesk_tickets=0,
        message=f"Currently analyzing reviews — 0 of {total_rows} processed",
    )

    for idx, row in enumerate(rows, start=1):
        parsed = _parse_review_row(row, default_rank=idx - 1)
        author = parsed["author"]
        text = parsed["text"]
        rating = parsed["rating"]
        play_review_id = parsed["play_review_id"]
        reviewed_at = parsed["reviewed_at"]
        play_rank = parsed["play_rank"]
        analyze_pct = 30 + int(((idx - 1) / max(1, total_rows)) * 35)
        _set_job(
            job_id,
            phase="analyzing",
            progress=min(65, analyze_pct),
            current_index=idx,
            total_reviews=total_rows,
            fetched=total_rows,
            processed=processed,
            new=new_count,
            refreshed=refreshed,
            skipped=skipped,
            jira_tickets=jira_tickets,
            zendesk_tickets=zendesk_tickets,
            message=(
                f"Currently analyzing review {idx} of {total_rows} "
                f"— detecting sentiment & category"
            ),
        )

        ok, reason, meta = _process_review(
            app_name,
            author,
            rating,
            text,
            batch_started_at=batch_started_at,
            play_review_id=play_review_id,
            reviewed_at=reviewed_at,
            play_rank=play_rank,
            skip_positive_tickets=skip_positive_tickets,
        )
        if ok:
            processed += 1
            if reason == "refreshed":
                refreshed += 1
            else:
                new_count += 1
            platform = meta.get("platform")
            sentiment = meta.get("sentiment", "")
            category = meta.get("category", "")
            if (
                reason == "processed"
                and skip_positive_tickets
                and sentiment == "positive"
                and not platform
            ):
                tickets_skipped_positive += 1
            if platform == "Jira":
                jira_tickets += 1
            elif platform == "Zendesk":
                zendesk_tickets += 1
            if reason == "processed" and platform:
                ticket_pct = 65 + int((idx / max(1, total_rows)) * 34)
                _set_job(
                    job_id,
                    phase="ticketing",
                    progress=min(99, ticket_pct),
                    current_index=idx,
                    processed=processed,
                    new=new_count,
                    refreshed=refreshed,
                    skipped=skipped,
                    jira_tickets=jira_tickets,
                    zendesk_tickets=zendesk_tickets,
                    message=(
                        f"Currently creating {platform} ticket for review "
                        f"{idx} of {total_rows}"
                    ),
                )
        else:
            skipped += 1

    db.session.commit()
    _set_job(
        job_id,
        status="completed",
        phase="finalize",
        progress=100,
        message=_completion_message(
            new_count,
            refreshed,
            skipped,
            tickets_created=jira_tickets + zendesk_tickets,
            tickets_skipped_positive=tickets_skipped_positive,
        ),
        batch_started_at=batch_started_at.isoformat(),
        processed=processed,
        new=new_count,
        refreshed=refreshed,
        skipped=skipped,
        fetched=total_rows,
        total_reviews=total_rows,
        current_index=total_rows,
        jira_tickets=jira_tickets,
        zendesk_tickets=zendesk_tickets,
        app_icon=app_icon or "https://cdn.simpleicons.org/googleplay/34A853",
        app_name=app_name,
    )


def _csv_export_rows(rows: list[Review]) -> list[list[str]]:
    out = []
    for idx, r in enumerate(rows, start=1):
        out.append(
            [
                str(idx),
                (r.app_name or "").strip(),
                (r.author or "").strip(),
                str(r.rating or ""),
                (r.sentiment or "").strip(),
                (r.category or "").strip(),
                str(r.confidence or ""),
                (r.content or "").replace("\n", " ").strip(),
                r.created_at.isoformat() if r.created_at else "",
            ]
        )
    return out


def _sentiment_style(sentiment: str) -> tuple[str, str]:
    s = (sentiment or "").lower()
    if s == "positive":
        return "DCFCE7", "166534"
    if s == "negative":
        return "FEE2E2", "991B1B"
    if s == "neutral":
        return "E2E8F0", "475569"
    return "EDE9FE", "5B21B6"


def _category_style(category: str) -> tuple[str, str]:
    c = (category or "").lower()
    if c == "support":
        return "DBEAFE", "1E40AF"
    if c == "complaint":
        return "FEF3C7", "92400E"
    if c == "bug":
        return "FFEDD5", "9A3412"
    if c == "feature_request":
        return "E0E7FF", "3730A3"
    return "F1F5F9", "475569"


def _professional_csv_bytes(report_title: str, rows: list[Review]) -> bytes:
    """UTF-8 BOM + structured sections for Excel/Sheets; plain CSV cannot store cell colors."""
    sentiment_counts = Counter([r.sentiment for r in rows])
    category_counts = Counter([r.category for r in rows])
    n = len(rows)
    avg_rating = round(sum(r.rating for r in rows) / n, 2) if n else 0
    total = max(1, n)

    buf = io.StringIO()
    writer = csv.writer(buf, dialect="excel", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(["ReviewBridge — export"])
    writer.writerow(["Report", report_title])
    writer.writerow(["Generated (UTC)", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
    writer.writerow(["Total rows", str(n)])
    writer.writerow(["Average rating", f"{avg_rating} / 5" if n else "—"])
    writer.writerow([])
    writer.writerow(["Sentiment", "Count", "Share %"])
    for key in ("positive", "negative", "neutral"):
        c = sentiment_counts.get(key, 0)
        writer.writerow([key, str(c), f"{round((c / total) * 100, 1)}"])
    other_sent = sum(v for k, v in sentiment_counts.items() if k not in ("positive", "negative", "neutral"))
    if other_sent:
        writer.writerow(["other", str(other_sent), f"{round((other_sent / total) * 100, 1)}"])
    writer.writerow([])
    writer.writerow(["Category", "Count", "Share %"])
    for key in ("bug", "feature_request", "support", "complaint"):
        c = category_counts.get(key, 0)
        label = key.replace("_", " ")
        writer.writerow([label, str(c), f"{round((c / total) * 100, 1)}"])
    writer.writerow([])
    writer.writerow(
        [
            "Note",
            "Use Excel export for colored cells matching the dashboard.",
        ]
    )
    writer.writerow([])
    writer.writerow(["— Data —"])
    writer.writerow(["sr_no", "app_name", "author", "rating", "sentiment", "category", "confidence", "review_text", "created_at"])
    writer.writerows(_csv_export_rows(rows))
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _xlsx_font_color(rgb6: str) -> str:
    rgb6 = (rgb6 or "").strip().lstrip("#").upper()
    if len(rgb6) == 6 and rgb6.isalnum():
        return f"FF{rgb6}"
    return "FF0F172A"


def _build_professional_xlsx(title: str, subtitle: str, rows: list[Review]) -> bytes:
    if not HAS_OPENPYXL:
        return b""

    sentiment_counts = Counter([r.sentiment for r in rows])
    category_counts = Counter([r.category for r in rows])
    n = len(rows)
    avg_rating = round(sum(r.rating for r in rows) / n, 2) if n else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, color="FFFFFF", size=13)
    thin = Side(style="thin", color="CBD5E1")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    last_col = 9
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = title_font
    c1.fill = header_fill
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(size=9, color=_xlsx_font_color("64748B"))
    c2.alignment = Alignment(horizontal="center", vertical="center")

    summary_rows = [
        ["Total reviews", n, "", "Positive", sentiment_counts.get("positive", 0), "", "Support", category_counts.get("support", 0)],
        ["Avg rating (mean)", f"{avg_rating} / 5" if n else "—", "", "Negative", sentiment_counts.get("negative", 0), "", "Complaint", category_counts.get("complaint", 0)],
        ["", "", "", "Neutral", sentiment_counts.get("neutral", 0), "", "Bug", category_counts.get("bug", 0)],
        ["", "", "", "", "", "", "Feature", category_counts.get("feature_request", 0)],
    ]
    for sr in summary_rows:
        ws.append(sr)
        rr = ws.max_row
        for cc in range(1, last_col + 1):
            ws.cell(row=rr, column=cc).border = grid
        ws.cell(row=rr, column=1).font = Font(bold=True)
        ws.cell(row=rr, column=4).font = Font(bold=True)
        ws.cell(row=rr, column=7).font = Font(bold=True)

    ws.append([])
    headers = ["#", "App", "Author", "Rating", "Sentiment", "Category", "Confidence", "Review", "Created (UTC)"]
    ws.append(headers)
    hr = ws.max_row
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=hr, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = grid
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, r in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                (r.app_name or "")[:80],
                (r.author or "")[:60],
                r.rating,
                (r.sentiment or "").title(),
                (r.category or "").replace("_", " ").title(),
                r.confidence,
                (r.content or "").replace("\n", " "),
                r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            ]
        )
        row_i = ws.max_row
        for col_idx in range(1, last_col + 1):
            ws.cell(row=row_i, column=col_idx).border = grid
            ws.cell(row=row_i, column=col_idx).alignment = wrap
        if (row_i - hr) % 2 == 0:
            zfill = PatternFill("solid", fgColor="F8FAFC")
            for col_idx in (1, 2, 3, 4, 7, 8, 9):
                ws.cell(row=row_i, column=col_idx).fill = zfill

        bg_s, fg_s = _sentiment_style(r.sentiment)
        s_cell = ws.cell(row=row_i, column=5)
        s_cell.fill = PatternFill("solid", fgColor=bg_s)
        s_cell.font = Font(bold=True, color=_xlsx_font_color(fg_s), size=9)

        bg_c, fg_c = _category_style(r.category)
        c_cell = ws.cell(row=row_i, column=6)
        c_cell.fill = PatternFill("solid", fgColor=bg_c)
        c_cell.font = Font(bold=True, color=_xlsx_font_color(fg_c), size=9)

    widths = {1: 5, 2: 22, 3: 16, 4: 8, 5: 12, 6: 14, 7: 10, 8: 52, 9: 18}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@main_bp.route("/")
def home():
    return render_template("home.html", popular_apps=POPULAR_APPS[:6])


@main_bp.route("/analysis")
def analysis():
    since = _dashboard_view_since()
    if since is None:
        session.pop("current_app_icon", None)

    if since is None:
        reviews = []
        positive_reviews = []
        non_positive_reviews = []
        tickets = []
        all_batch_tickets = []
        logs = []
        current_app_label = "No dataset loaded"
        current_app_icon = "https://cdn.simpleicons.org/googleplay/34A853"
    else:
        reviews = _batch_reviews_query(since).all()
        positive_reviews = [r for r in reviews if r.sentiment == "positive"]
        non_positive_reviews = [r for r in reviews if r.sentiment != "positive"]
        app_names = sorted({r.app_name for r in reviews if r.app_name})
        current_app_label = app_names[0] if len(app_names) == 1 else ", ".join(app_names[:3])
        if len(app_names) > 3:
            current_app_label += "..."
        current_app_icon = (
            POPULAR_APPS_BY_NAME.get(app_names[0], {}).get("icon")
            if len(app_names) == 1
            else "https://cdn.simpleicons.org/googleplay/34A853"
        )
        all_batch_tickets = (
            Ticket.query.join(Review, Ticket.review_id == Review.id)
            .filter(_batch_review_filter(since))
            .order_by(Ticket.created_at.desc())
            .all()
        )
        tickets = all_batch_tickets
        logs = ProcessingLog.query.filter(ProcessingLog.created_at >= since).order_by(ProcessingLog.created_at.desc()).limit(10).all()

    total_reviews = len(reviews)
    avg_rating = round(sum([r.rating for r in reviews]) / total_reviews, 2) if total_reviews else 0
    sentiment_counts = Counter([r.sentiment for r in reviews])
    category_counts = Counter([r.category for r in reviews])
    ticket_counts = Counter([t.platform for t in all_batch_tickets])
    total = max(1, total_reviews)
    percentages = {
        "positive": round((sentiment_counts.get("positive", 0) / total) * 100, 1) if total_reviews else 0,
        "negative": round((sentiment_counts.get("negative", 0) / total) * 100, 1) if total_reviews else 0,
        "neutral": round((sentiment_counts.get("neutral", 0) / total) * 100, 1) if total_reviews else 0,
        "complaint": round((category_counts.get("complaint", 0) / total) * 100, 1) if total_reviews else 0,
        "support": round((category_counts.get("support", 0) / total) * 100, 1) if total_reviews else 0,
    }

    active_id = session.get("active_fetch_job_id")
    if active_id and not _get_job(active_id):
        session.pop("active_fetch_job_id", None)
    active_fetch_job_id = session.get("active_fetch_job_id")

    has_review_results = since is not None and total_reviews > 0
    if not has_review_results:
        session.pop("pipeline_snapshot", None)
    pipeline_snapshot = session.get("pipeline_snapshot") if has_review_results else None
    show_pipeline_card = True

    return render_template(
        "analysis.html",
        reviews=reviews,
        positive_reviews=positive_reviews[:12],
        non_positive_reviews=non_positive_reviews[:12],
        tickets=tickets,
        logs=logs,
        popular_apps=POPULAR_APPS,
        is_clean=(since is None),
        current_app_label=current_app_label,
        current_app_icon=current_app_icon,
        active_context_title="Live focus",
        active_fetch_job_id=active_fetch_job_id,
        has_review_results=has_review_results,
        show_pipeline_card=show_pipeline_card,
        pipeline_snapshot=pipeline_snapshot,
        dashboard_since_iso=(since.isoformat() if since else None),
        has_dashboard_export_data=(total_reviews > 0),
        percentages=percentages,
        stats={
            "total_reviews": total_reviews,
            "avg_rating": avg_rating,
            "positive": sentiment_counts.get("positive", 0),
            "negative": sentiment_counts.get("negative", 0),
            "neutral": sentiment_counts.get("neutral", 0),
            "bugs": category_counts.get("bug", 0),
            "features": category_counts.get("feature_request", 0),
            "support": category_counts.get("support", 0),
            "complaints": category_counts.get("complaint", 0),
            "jira": ticket_counts.get("Jira", 0),
            "zendesk": ticket_counts.get("Zendesk", 0),
            "tickets_total": len(all_batch_tickets),
        },
    )


@main_bp.route("/dashboard")
def dashboard_redirect():
    since = request.args.get("since")
    if since:
        return redirect(url_for("main.analysis", since=since))
    return redirect(url_for("main.analysis"))


@main_bp.route("/api/storage-health")
def api_storage_health():
    """JSON diagnostics: reviews vs tickets and duplicate indicators."""
    return jsonify(storage_health_report())


@main_bp.route("/history")
def history():
    reviews = Review.query.order_by(Review.created_at.desc()).limit(5000).all()
    tickets = (
        db.session.query(Ticket, Review.app_name)
        .join(Review, Ticket.review_id == Review.id)
        .order_by(Ticket.created_at.desc())
        .limit(2000)
        .all()
    )
    logs = ProcessingLog.query.order_by(ProcessingLog.created_at.desc()).limit(200).all()

    grouped = {}
    for review in reviews:
        app_name = review.app_name or "Unknown App"
        grouped.setdefault(app_name, {"reviews": [], "tickets": []})
        grouped[app_name]["reviews"].append(review)

    for ticket, app_name in tickets:
        app_key = app_name or "Unknown App"
        grouped.setdefault(app_key, {"reviews": [], "tickets": []})
        grouped[app_key]["tickets"].append(ticket)

    app_history = [{"app_name": name, **data} for name, data in grouped.items()]
    for block in app_history:
        block["reviews"].sort(key=lambda r: r.created_at or datetime.min, reverse=True)
        block["tickets"].sort(key=lambda t: t.created_at or datetime.min, reverse=True)

    app_history.sort(
        key=lambda item: max((r.created_at for r in item["reviews"]), default=datetime.min),
        reverse=True,
    )

    return render_template(
        "history.html",
        app_history=app_history,
        logs=logs,
        has_history_export_data=(len(reviews) > 0),
        popular_apps_by_name=POPULAR_APPS_BY_NAME,
    )


@main_bp.route("/history/clear", methods=["POST"])
def clear_history():
    session.pop("active_fetch_job_id", None)
    session.pop("pipeline_snapshot", None)
    Ticket.query.delete()
    Review.query.delete()
    ProcessingLog.query.delete()
    db.session.commit()
    log_message("History cleared: all reviews, tickets, and logs removed.", "info")
    flash("History cleared. All saved reviews, tickets, and logs were removed.", "success")
    return redirect(url_for("main.history"))


@main_bp.route("/dashboard/clear", methods=["POST"])
def clear_dashboard():
    session.pop("active_fetch_job_id", None)
    session.pop("pipeline_snapshot", None)
    session.pop("current_app_icon", None)
    flash("Dashboard cleared. Run a new fetch/upload to show latest batch.", "info")
    return redirect(url_for("main.analysis"))


@main_bp.route("/export/dashboard.csv")
def export_dashboard_csv():
    since = _dashboard_view_since()
    if since is None:
        rows = []
    else:
        rows = _batch_reviews_query(since).all()

    data = _professional_csv_bytes("Dashboard Latest Reviews", rows)
    return Response(
        data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=dashboard_latest_reviews.csv"},
    )


@main_bp.route("/export/history.csv")
def export_history_csv():
    rows = Review.query.order_by(Review.created_at.desc()).limit(2000).all()
    data = _professional_csv_bytes("History Reviews", rows)
    return Response(
        data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=history_reviews.csv"},
    )


@main_bp.route("/export/dashboard.xlsx")
def export_dashboard_xlsx():
    since = _dashboard_view_since()
    if since is None:
        rows = []
    else:
        rows = _batch_reviews_query(since).all()
    xlsx = _build_professional_xlsx(
        "Dashboard Latest Reviews",
        f"Generated UTC {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} · {len(rows)} row(s)",
        rows,
    )
    if not xlsx:
        flash("Excel export needs the openpyxl package. Run: pip install -r requirements.txt", "danger")
        q_since = request.args.get("since")
        if q_since:
            return redirect(url_for("main.analysis", since=q_since))
        return redirect(url_for("main.analysis"))
    return Response(
        xlsx,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dashboard_latest_reviews.xlsx"},
    )


@main_bp.route("/export/history.xlsx")
def export_history_xlsx():
    rows = Review.query.order_by(Review.created_at.desc()).limit(2000).all()
    xlsx = _build_professional_xlsx(
        "History Reviews",
        f"Generated UTC {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} · {len(rows)} row(s)",
        rows,
    )
    if not xlsx:
        flash("Excel export needs the openpyxl package. Run: pip install -r requirements.txt", "danger")
        return redirect(url_for("main.history"))
    return Response(
        xlsx,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=history_reviews.xlsx"},
    )


@main_bp.route("/api/app-catalog/status")
def app_catalog_status():
    return jsonify(catalog_status())


@main_bp.route("/api/app-catalog")
def app_catalog():
    return jsonify(load_catalog())


@main_bp.route("/api/app-suggestions")
def app_suggestions():
    query = (request.args.get("q") or "").strip()
    lang = (request.args.get("lang") or "").strip()
    country = (request.args.get("country") or "us").strip()
    limit = max(1, min(50, int(request.args.get("limit") or 12)))

    if len(query) < 1:
        return jsonify([])

    try:
        if is_package_query(query):
            local = lookup_local_by_package(query)
            if local:
                return jsonify([local])
            play_pkg = lookup_app_by_package(query, lang=lang, country=country)
            if play_pkg:
                return jsonify([play_pkg])
            return jsonify([])

        local = search_local_catalog(query, limit=limit)
        need_play = (
            len(local) == 0
            or len(local) < min(3, limit)
            or not has_strong_local_match(query, local)
        )
        play: list = []
        if need_play:
            play = search_apps_play(query=query, limit=limit, lang=lang, country=country)
        apps = merge_and_rank_suggestions(local, play, query, limit)
    except Exception as e:
        current_app.logger.exception("App suggestion search failed: %s", e)
        return jsonify([])

    return jsonify(apps)


def _run_fetch_job(
    app,
    job_id: str,
    package_name: str,
    app_name: str,
    count: int,
    lang: str,
    country: str,
    sort: str,
    app_icon: str,
    fetch_all: bool,
    skip_positive_tickets: bool = False,
):
    with app.app_context():
        try:
            _set_job(
                job_id,
                status="running",
                job_type="play_fetch",
                phase="prepare",
                progress=2,
                app_name=app_name,
                app_icon=app_icon,
                message="Currently connecting to Google Play…",
            )
            def _fetch_progress(fetched: int, total_goal: int | None, meta: dict | None = None):
                meta = meta or {}
                cc = meta.get("country") or ""
                if fetch_all or not total_goal:
                    pct = min(28, 5 + min(23, int(fetched / 120)))
                    msg = (
                        f"Currently downloading reviews ({cc or 'storefront'}) — "
                        f"{fetched} collected so far…"
                    )
                else:
                    pct = min(28, 5 + int((fetched / max(1, total_goal)) * 23))
                    msg = (
                        f"Currently downloading reviews ({cc}) — "
                        f"{fetched} of {total_goal} unique reviews…"
                    )
                _set_job(
                    job_id,
                    phase="load",
                    progress=pct,
                    fetched=fetched,
                    total_reviews=total_goal or 0,
                    message=msg,
                )
            if fetch_all:
                rows = fetch_google_play_reviews_all(
                    package_name=package_name,
                    lang=lang,
                    country=country,
                    sort=sort,
                    progress_callback=_fetch_progress,
                )
            else:
                rows = fetch_google_play_reviews(
                    package_name=package_name,
                    count=count,
                    lang=lang,
                    country=country,
                    sort=sort,
                    progress_callback=_fetch_progress,
                )

            total_rows = len(rows)
            batch_started_at = _batch_now()

            if total_rows > 0:
                _set_job(
                    job_id,
                    phase="load",
                    progress=28,
                    fetched=total_rows,
                    total_reviews=total_rows,
                    message=f"Downloaded {total_rows} reviews — starting analysis…",
                )
                _process_reviews_loop(
                    job_id,
                    app_name,
                    rows,
                    batch_started_at,
                    app_icon,
                    skip_positive_tickets=skip_positive_tickets,
                )
            if (_get_job(job_id) or {}).get("status") == "completed":
                job_done = _get_job(job_id) or {}
                log_message(
                    f"Google Play fetch complete for {package_name}: "
                    f"{job_done.get('new', 0)} new, {job_done.get('refreshed', 0)} refreshed, "
                    f"{job_done.get('skipped', 0)} skipped.",
                    "info",
                )
        except Exception as e:
            db.session.rollback()
            log_message(f"Google Play fetch failed: {e}", "error")
            _set_job(
                job_id,
                status="error",
                phase="finalize",
                progress=100,
                message=str(e),
            )


def _run_csv_job(
    app,
    job_id: str,
    app_name: str,
    rows: list,
    app_icon: str = "",
    skip_positive_tickets: bool = False,
):
    with app.app_context():
        try:
            _set_job(
                job_id,
                status="running",
                job_type="csv_upload",
                phase="prepare",
                progress=5,
                app_name=app_name,
                app_icon=app_icon or "https://cdn.simpleicons.org/googledocs/4285F4",
                message="Currently reading CSV file…",
            )

            total_rows = len(rows)
            batch_started_at = _batch_now()
            _set_job(
                job_id,
                phase="load",
                progress=25,
                fetched=total_rows,
                total_reviews=total_rows,
                message=f"Parsed {total_rows} rows from CSV — starting analysis…",
            )

            _process_reviews_loop(
                job_id,
                app_name,
                rows,
                batch_started_at,
                app_icon,
                skip_positive_tickets=skip_positive_tickets,
            )
            if (_get_job(job_id) or {}).get("status") == "completed":
                job_done = _get_job(job_id) or {}
                log_message(
                    f"CSV upload complete: "
                    f"{job_done.get('new', 0)} new, {job_done.get('refreshed', 0)} refreshed, "
                    f"{job_done.get('skipped', 0)} skipped.",
                    "info",
                )
        except Exception as e:
            db.session.rollback()
            log_message(f"CSV upload failed: {e}", "error")
            _set_job(
                job_id,
                status="error",
                phase="finalize",
                progress=100,
                message=str(e),
            )


@main_bp.route("/upload/start", methods=["POST"])
def start_csv_upload():
    file = request.files.get("review_file")
    app_name = (request.form.get("app_name") or "Unknown App").strip()

    if not file or file.filename == "":
        return jsonify({"ok": False, "error": "Please upload a CSV file."}), 400

    try:
        content = file.read().decode("utf-8")
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = []
        for row in csv_reader:
            entry = {
                "author": (row.get("author") or "anonymous").strip(),
                "content": (row.get("content") or "").strip(),
                "rating": int(row.get("rating") or 3),
            }
            for key in ("at", "reviewed_at", "date", "review_date"):
                if row.get(key):
                    entry[key] = row.get(key)
            rows.append(entry)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Could not parse CSV: {e}"}), 400

    if not rows:
        return jsonify({"ok": False, "error": "CSV file has no data rows."}), 400

    skip_positive_tickets = _parse_skip_positive_tickets(request.form.get("skip_positive_tickets"))

    job_id = str(uuid.uuid4())
    session["active_fetch_job_id"] = job_id
    session.pop("pipeline_snapshot", None)
    app_icon = "https://cdn.simpleicons.org/googledocs/4285F4"
    _set_job(
        job_id,
        status="queued",
        job_type="csv_upload",
        phase="prepare",
        progress=0,
        app_name=app_name,
        app_icon=app_icon,
        message="Starting CSV analysis pipeline…",
        fetched=0,
        processed=0,
        new=0,
        refreshed=0,
        skipped=0,
        jira_tickets=0,
        zendesk_tickets=0,
    )

    app = current_app._get_current_object()
    worker = threading.Thread(
        target=_run_csv_job,
        args=(app, job_id, app_name, rows, app_icon, skip_positive_tickets),
        daemon=True,
    )
    worker.start()

    return jsonify({"ok": True, "job_id": job_id})


@main_bp.route("/upload", methods=["POST"])
def upload_reviews():
    file = request.files.get("review_file")
    app_name = request.form.get("app_name", "Unknown App").strip()

    if not file or file.filename == "":
        flash("Please upload a CSV file.", "danger")
        return redirect(url_for("main.analysis"))

    content = file.read().decode("utf-8")
    csv_reader = csv.DictReader(io.StringIO(content))

    skip_positive_tickets = _parse_skip_positive_tickets(request.form.get("skip_positive_tickets"))
    batch_started_at = _batch_now()
    processed = 0
    skipped = 0

    for rank, row in enumerate(csv_reader):
        parsed = _parse_review_row(row, default_rank=rank)
        ok, _, _ = _process_review(
            app_name,
            parsed["author"],
            parsed["rating"],
            parsed["text"],
            batch_started_at=batch_started_at,
            play_review_id=parsed["play_review_id"],
            reviewed_at=parsed["reviewed_at"],
            play_rank=parsed["play_rank"],
            skip_positive_tickets=skip_positive_tickets,
        )
        if ok:
            processed += 1
        else:
            skipped += 1

    db.session.commit()
    log_message(f"CSV upload complete: {processed} processed, {skipped} skipped.", "info")

    flash(
        f"Processing done. {processed} reviews in batch, {skipped} skipped (invalid only).",
        "success",
    )
    return redirect(url_for("main.analysis", since=batch_started_at.isoformat()))


@main_bp.route("/fetch", methods=["POST"])
def fetch_from_google_play():
    package_name = (request.form.get("package_name") or "").strip()
    app_name = (request.form.get("app_name") or package_name or "Unknown App").strip()
    lang = (request.form.get("lang") or "").strip()
    country = (request.form.get("country") or "us").strip()
    sort = (request.form.get("sort") or "newest").strip()
    fetch_all = request.form.get("fetch_all") in ("1", "true", "on", "yes")
    skip_positive_tickets = _parse_skip_positive_tickets(request.form.get("skip_positive_tickets"))

    if not package_name:
        flash("Package name is required. Example: com.whatsapp", "danger")
        return redirect(url_for("main.analysis"))

    try:
        count = _parse_review_count(request.form.get("count"))
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("main.analysis"))

    try:
        if fetch_all:
            rows = fetch_google_play_reviews_all(
                package_name=package_name,
                lang=lang,
                country=country,
                sort=sort,
            )
        else:
            rows = fetch_google_play_reviews(
                package_name=package_name,
                count=count,
                lang=lang,
                country=country,
                sort=sort,
            )
    except Exception as e:
        log_message(f"Google Play fetch failed: {e}", "error")
        flash(f"Google Play fetch failed: {e}", "danger")
        return redirect(url_for("main.analysis"))

    batch_started_at = _batch_now()
    processed = 0
    skipped = 0

    for rank, row in enumerate(rows):
        parsed = _parse_review_row(row, default_rank=rank)
        ok, _, _ = _process_review(
            app_name,
            parsed["author"],
            parsed["rating"],
            parsed["text"],
            batch_started_at=batch_started_at,
            play_review_id=parsed["play_review_id"],
            reviewed_at=parsed["reviewed_at"],
            play_rank=parsed["play_rank"],
            skip_positive_tickets=skip_positive_tickets,
        )
        if ok:
            processed += 1
        else:
            skipped += 1

    db.session.commit()
    log_message(
        f"Google Play fetch complete for {package_name}: {processed} processed, {skipped} skipped.",
        "info",
    )

    flash(
        f"Fetched {len(rows)} reviews from Google Play. {processed} in batch, {skipped} skipped (invalid only).",
        "success",
    )
    return redirect(url_for("main.analysis", since=batch_started_at.isoformat()))


@main_bp.route("/fetch/start", methods=["POST"])
def start_fetch_from_google_play():
    package_name = (request.form.get("package_name") or "").strip()
    app_name = (request.form.get("app_name") or package_name or "Unknown App").strip()
    lang = (request.form.get("lang") or "").strip()
    country = (request.form.get("country") or "us").strip()
    sort = (request.form.get("sort") or "newest").strip()
    app_icon = (request.form.get("app_icon") or "").strip() or "https://cdn.simpleicons.org/googleplay/34A853"
    fetch_all = request.form.get("fetch_all") in ("1", "true", "on", "yes")
    skip_positive_tickets = _parse_skip_positive_tickets(request.form.get("skip_positive_tickets"))

    if not package_name:
        return jsonify({"ok": False, "error": "Package name is required."}), 400

    try:
        count = _parse_review_count(request.form.get("count"))
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    job_id = str(uuid.uuid4())
    session["active_fetch_job_id"] = job_id
    session.pop("pipeline_snapshot", None)
    _set_job(
        job_id,
        status="queued",
        job_type="play_fetch",
        phase="prepare",
        progress=0,
        app_name=app_name,
        app_icon=app_icon,
        message="Starting analysis pipeline…",
        fetched=0,
        processed=0,
        new=0,
        refreshed=0,
        skipped=0,
        jira_tickets=0,
        zendesk_tickets=0,
    )

    app = current_app._get_current_object()
    worker = threading.Thread(
        target=_run_fetch_job,
        args=(
            app,
            job_id,
            package_name,
            app_name,
            count,
            lang,
            country,
            sort,
            app_icon,
            fetch_all,
            skip_positive_tickets,
        ),
        daemon=True,
    )
    worker.start()

    return jsonify({"ok": True, "job_id": job_id})


@main_bp.route("/fetch/status/<job_id>")
def fetch_status(job_id: str):
    job = _get_job(job_id)
    if not job:
        return jsonify({"ok": False, "error": "Job not found", "stale": True}), 404
    return jsonify({"ok": True, **job})


@main_bp.route("/fetch/dismiss-active", methods=["POST"])
def dismiss_active_fetch_job():
    session.pop("active_fetch_job_id", None)
    return jsonify({"ok": True})


@main_bp.route("/fetch/activate/<job_id>", methods=["POST"])
def activate_fetch_batch(job_id: str):
    job = _get_job(job_id)
    if not job or job.get("status") != "completed":
        return jsonify({"ok": False, "error": "Job is not ready to activate"}), 400

    batch_since = job.get("batch_started_at")
    if not batch_since:
        return jsonify({"ok": False, "error": "Job has no batch timestamp"}), 400

    snapshot = _pipeline_snapshot_from_job(job)
    session["pipeline_snapshot"] = snapshot
    session.pop("active_fetch_job_id", None)
    return jsonify(
        {
            "ok": True,
            "batch_started_at": batch_since,
            "app_icon": job.get("app_icon") or "https://cdn.simpleicons.org/googleplay/34A853",
            "pipeline_snapshot": snapshot,
        }
    )


